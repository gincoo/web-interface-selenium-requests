# coding=utf8
import openpyxl
import sys
from collections.abc import Iterable
import os

base_path = os.path.dirname(os.getcwd())


#
# Excel 文件操作类
# 封装 openpyxl 框架,操作.xlsx 文件
#
class HandExcel:

    def __init__(self):
        self.excel_handle = openpyxl.load_workbook(os.path.abspath(base_path + "/Case/imooc.xlsx"))

    # def load_excel(self):
    #     '''
    #     加载excel
    #     '''
    #     open_excel = openpyxl.load_workbook(os.path.abspath(base_path + "/Case/imooc.xlsx"))
    #     return open_excel

    def get_sheet_data(self, index=None):
        """
        加载对应sheet内的所有内容,
        :param index: sheet角标,不传默认为第一个sheet
        :return: 返回
        """
        sheet_names = self.excel_handle.sheetnames
        if index == None:
            index = 0
        #key为sheetname
        data = self.excel_handle[sheet_names[index]]
        return data

    def get_cell_value(self, row, cols):
        """
        获取目标单元格内容
        :param row: 行
        :param cols: 列
        :return: 返回具体内容
        """
        data = self.get_sheet_data().cell(row=row, column=cols).value
        return data

    def get_rows(self):
        """
        获取最大行数,默认为第一个sheet
        :return: 最大行数
        """
        row = self.get_sheet_data().max_row
        return row

    def get_rows_value(self, row):
        """
        获取某一行的内容
        :param row: 行数
        :return: 返回某一行的所有内容
        """
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    def excel_write_data(self, row, cols, value):
        """
        写入数据
        :param row: 行
        :param cols:列
        :param value: 写入的数据
        """
        wb = self.excel_handle
        wr = wb.active
        wr.cell(row, cols, value)
        wb.save(base_path + "/Case/imooc.xlsx")

    def get_columns_value(self, key=None):
        """
        获取某一列得数据
        :param key: 目标列
        :return: 返回目标列的数据数组
        """
        columns_list = []
        if key == None:
            key = 'A'
        # 对应列下的所有数据对象
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id):
        """
        获取对应 caseid 的行号
        :param case_id: caseid
        :return: 返回行号
        """
        num = 1
        cols_data = self.get_columns_value()
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return num

    def get_excel_data(self):
        """
        获取excel里面所有的数据
        :return: 获取
        """
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i + 2))

        return data_list


excel_data = HandExcel()

if __name__ == "__main__":
    handle = HandExcel()
    ##print(handle.get_rows_number('imooc_001'))
    print(handle.get_excel_data())
